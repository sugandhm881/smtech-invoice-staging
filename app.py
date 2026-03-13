import os
import io
import json
import logging
import base64
import zipfile
import tempfile
import smtplib
import random
import string
import qrcode
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import date, datetime, timedelta, timezone
from urllib.parse import unquote
from werkzeug.security import generate_password_hash, check_password_hash

from openpyxl import Workbook
from flask import Flask, request, send_file, jsonify, render_template, redirect, url_for, flash, session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from dotenv import load_dotenv
from fpdf import FPDF
from PIL import Image

# --- SUPABASE SETUP ---
from supabase import create_client, Client
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

supabase: Client = None
if SUPABASE_URL and SUPABASE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
else:
    logging.warning("No Supabase Credentials found. DB calls will fail.")

# ------------------ CONFIG ------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CALIBRI_FONT_PATH = os.path.join(BASE_DIR, "CALIBRI.TTF")
DEFAULT_LOGO = os.path.join(BASE_DIR, "static", "logo.png") 
DEFAULT_SIGNATURE = os.path.join(BASE_DIR, "static", "Signatory.png")
TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=os.path.join(BASE_DIR, "static"))

_secret = os.getenv("SECRET_KEY", "fallback-secret-for-staging")
app.secret_key = _secret

from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "60 per hour"],
    storage_uri="memory://"
)

EMAIL_HOST = os.getenv('EMAIL_HOST', 'smtp.gmail.com')
EMAIL_PORT = int(os.getenv('EMAIL_PORT', 587))
EMAIL_USER = os.getenv('EMAIL_USER')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD')

UPI_ID = "sugandh.mishra1@ybl"
UPI_NAME = "SM Tech"
REPORT_HOUR_UTC = 16 

STATE_CODES = {
    "Jammu and Kashmir": "01", "Himachal Pradesh": "02", "Punjab": "03", "Chandigarh": "04",
    "Uttarakhand": "05", "Haryana": "06", "Delhi": "07", "Rajasthan": "08", "Uttar Pradesh": "09",
    "Bihar": "10", "Sikkim": "11", "Arunachal Pradesh": "12", "Nagaland": "13", "Manipur": "14",
    "Mizoram": "15", "Tripura": "16", "Meghalaya": "17", "Assam": "18", "West Bengal": "19",
    "Jharkhand": "20", "Odisha": "21", "Chhattisgarh": "22", "Madhya Pradesh": "23",
    "Gujarat": "24", "Dadra and Nagar Haveli": "26", "Maharashtra": "27", "Karnataka": "29",
    "Goa": "30", "Lakshadweep": "31", "Kerala": "32", "Tamil Nadu": "33", "Puducherry": "34",
    "Andaman and Nicobar Islands": "35", "Telangana": "36", "Andhra Pradesh": "37", "Ladakh": "38"
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ------------------ AUTH ------------------
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

MASTER_USERNAME = os.getenv("LOGIN_USER", "admin")
MASTER_PASSWORD = os.getenv("LOGIN_PASS", "password")

class User(UserMixin):
    def __init__(self, id, is_master=False, payment_active=True, permissions=None):
        self.id = id
        self.is_master = is_master
        self.payment_active = payment_active
        self.permissions = permissions if permissions is not None else ['sale', 'purchase']

    @property
    def is_active(self):
        return True
    
    def has_permission(self, perm):
        if self.is_master: return True
        return perm in self.permissions

@login_manager.user_loader
def load_user(user_id):
    # Retrieve updated master username from configs if it was renamed
    master_id = MASTER_USERNAME
    try:
        mc = supabase.table('configs').select('profile').eq('tenant_id', 'master_config').execute()
        if mc.data and mc.data[0].get('profile') and 'master_username' in mc.data[0]['profile']:
            master_id = mc.data[0]['profile']['master_username']
    except: pass

    if user_id == master_id:
        return User(user_id, is_master=True, payment_active=True, permissions=['sale', 'purchase'])
    
    try:
        user_doc = supabase.table('app_users').select('*').eq('username', user_id).execute()
        if user_doc.data:
            data = user_doc.data[0]
            return User(user_id, is_master=False, payment_active=data.get('is_active', False), permissions=data.get('permissions'))
    except: pass
    return None

@app.before_request
def check_activation():
    if current_user.is_authenticated:
        if not current_user.is_master and not current_user.payment_active:
            if request.endpoint not in ['activation_page', 'logout', 'static', 'check_status_api']:
                return redirect(url_for('activation_page'))

# ------------------ HELPERS ------------------
def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

def send_email_raw(to_email, subject, body):
    try:
        msg = MIMEText(body)
        msg['Subject'] = subject
        msg['From'] = EMAIL_USER
        msg['To'] = to_email
        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASSWORD)
            server.send_message(msg)
        return True
    except Exception as e:
        logging.error(f"Email Error: {e}")
        return False

def compress_image(file_storage, max_width=400):
    try:
        img = Image.open(file_storage)
        width_percent = (max_width / float(img.size[0]))
        if width_percent < 1:
            new_height = int((float(img.size[1]) * float(width_percent)))
            img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
        if img.mode not in ('RGBA', 'RGB', 'L'):
            img = img.convert('RGBA')
        output = io.BytesIO()
        img.save(output, format='PNG', optimize=True)
        return base64.b64encode(output.getvalue()).decode('utf-8')
    except Exception as e:
        return None

def generate_upi_qr_base64(upi_id, upi_name, amount):
    try:
        upi_str = f"upi://pay?pa={upi_id}&pn={upi_name}&am={amount:.2f}&cu=INR"
        qr = qrcode.QRCode(version=1, box_size=4, border=2)
        qr.add_data(upi_str)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return base64.b64encode(buf.getvalue()).decode('utf-8')
    except Exception as e:
        return None

# ------------------ DATABASE TENANT HELPER ------------------
def get_tenant_id(target_user=None):
    """Determines the active tenant (schema boundary) for queries"""
    master_id = session.get('master_id', MASTER_USERNAME)
    if target_user:
        return 'master' if target_user == master_id else target_user
    if current_user.is_authenticated and not current_user.is_master:
        return current_user.id
    view_mode = session.get('view_mode')
    if current_user.is_authenticated and current_user.is_master and view_mode and view_mode != master_id:
        return view_mode
    return 'master'

def get_all_users():
    master_id = session.get('master_id', MASTER_USERNAME)
    users = [master_id]
    try:
        res = supabase.table('app_users').select('username').execute()
        for doc in res.data:
            users.append(doc['username'])
    except: pass
    return sorted(users)

@app.context_processor
def inject_global_data():
    if current_user.is_authenticated:
        profile = get_seller_profile_data()
        all_users = get_all_users() if current_user.is_master else []
        viewing_user = session.get('view_mode', current_user.id)
        return dict(profile=profile, current_user=current_user, all_users=all_users, viewing_user=viewing_user)
    return dict(profile={}, current_user=None, all_users=[], viewing_user=None)

# ------------------ SUPABASE DB HELPERS ------------------
def get_seller_profile_data(target_user_id=None):
    tenant = get_tenant_id(target_user_id)
    try:
        res = supabase.table('configs').select('profile').eq('tenant_id', tenant).execute()
        if res.data and res.data[0].get('profile'):
            return res.data[0]['profile']
    except Exception as e:
        logging.error(f"Profile error: {e}")
    return {"company_name": "Sahayak ERP", "invoice_prefix": "SHK"}

def save_seller_profile_data(data, target_user_id=None):
    tenant = get_tenant_id(target_user_id)
    res = supabase.table('configs').select('tenant_id').eq('tenant_id', tenant).execute()
    if res.data:
        supabase.table('configs').update({'profile': data}).eq('tenant_id', tenant).execute()
    else:
        supabase.table('configs').insert({'tenant_id': tenant, 'profile': data, 'counters': {}}).execute()

def load_clients():
    tenant = get_tenant_id()
    res = supabase.table('clients').select('name, data').eq('tenant_id', tenant).execute()
    return {r['name']: r['data'] for r in res.data}

def save_single_client(name, data):
    tenant = get_tenant_id()
    supabase.table('clients').upsert({'tenant_id': tenant, 'name': name, 'data': data}).execute()

def load_particulars():
    tenant = get_tenant_id()
    res = supabase.table('particulars').select('name, data').eq('tenant_id', tenant).execute()
    return {r['name']: r['data'] for r in res.data}

def save_single_particular(name, data):
    tenant = get_tenant_id()
    supabase.table('particulars').upsert({'tenant_id': tenant, 'name': name, 'data': data}).execute()

def get_collection_name(data):
    cat = data.get('doc_category', 'sale')
    dtype = data.get('doc_type', 'invoice')
    is_cn = data.get('is_credit_note', False)
    is_dn = data.get('is_debit_note', False)
    if cat == 'purchase':
        if is_dn: return 'purchase_debit_notes'
        if dtype == 'po': return 'purchase_orders'
        if dtype == 'grn': return 'purchase_grns'
        if dtype == 'bill': return 'purchase_bills'
        return 'purchase_misc'
    else:
        if is_cn: return 'sales_credit_notes'
        if is_dn: return 'sales_debit_notes'
        return 'sales_invoices'

def load_invoices_for_user(target_user_id):
    tenant = get_tenant_id(target_user_id)
    res = supabase.table('documents').select('data').eq('tenant_id', tenant).execute()
    return [r['data'] for r in res.data]

def load_invoices():
    tenant = get_tenant_id()
    res = supabase.table('documents').select('data').eq('tenant_id', tenant).execute()
    return [r['data'] for r in res.data]

def save_single_invoice(invoice_data):
    tenant = get_tenant_id()
    coll = get_collection_name(invoice_data)
    bill_no = invoice_data['bill_no'].replace('/', '_')
    supabase.table('documents').upsert({
        'tenant_id': tenant,
        'bill_no': bill_no,
        'collection_name': coll,
        'data': invoice_data
    }).execute()

def get_next_counter(is_credit_note=False, is_debit_note=False, is_purchase=False, doc_type='invoice'):
    tenant = get_tenant_id()
    if is_purchase:
        if is_debit_note: field = "pdn_counter"
        elif doc_type == 'po': field = "po_counter"
        elif doc_type == 'grn': field = "grn_counter"
        elif doc_type == 'bill': field = "pb_counter"
        else: field = "po_counter" 
    else:
        if is_credit_note: field = "cn_counter"
        elif is_debit_note: field = "dn_counter"
        else: field = "counter" 
        
    res = supabase.rpc('increment_counter', {'p_tenant_id': tenant, 'p_field': field}).execute()
    return res.data

def get_all_activation_requests():
    try:
        res = supabase.table('activation_requests').select('data').order('created_at', desc=True).execute()
        return [r['data'] for r in res.data]
    except: return []

def send_email_with_attachment(to_email, subject, body, attachment_bytes, filename):
    msg = MIMEMultipart()
    msg['From'] = EMAIL_USER
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment_bytes.getvalue())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={filename}')
    msg.attach(part)
    with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        server.send_message(msg)

def convert_to_words(number):
    units = ["","One","Two","Three","Four","Five","Six","Seven","Eight","Nine","Ten","Eleven","Twelve","Thirteen","Fourteen","Fifteen","Sixteen","Seventeen","Eighteen","Nineteen"]
    tens = ["","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"]
    def two_digit(n): return units[n] if n < 20 else tens[n//10] + (" " + units[n%10] if n%10 else "")
    def three_digit(n):
        s = ""
        if n >= 100: s += units[n//100] + " Hundred" + (" " if n % 100 else "")
        if n % 100: s += two_digit(n%100)
        return s
    n = int(abs(number))
    paise = round((abs(number) - n) * 100)
    crore = n // 10000000; n %= 10000000
    lakh = n // 100000; n %= 100000
    thousand = n // 1000; n %= 1000
    hundred = n
    parts = []
    if crore: parts.append(three_digit(crore) + " Crore")
    if lakh: parts.append(three_digit(lakh) + " Lakh")
    if thousand: parts.append(three_digit(thousand) + " Thousand")
    if hundred: parts.append(three_digit(hundred))
    words = " ".join(parts) if parts else "Zero"
    if paise: words += f" and {two_digit(paise)} Paise"
    return words + " Only"

def PDF_Generator(invoice_data, is_credit_note=False, is_debit_note=False):
    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("Calibri", "", CALIBRI_FONT_PATH, uni=True)
    pdf.add_font("Calibri", "B", CALIBRI_FONT_PATH, uni=True)

    profile = get_seller_profile_data()
    margin, page_width, line_height = 15, pdf.w - 30, 5 
    col_width = (page_width / 2) - 5 
    
    invoice_type = invoice_data.get('invoice_type', profile.get('invoice_type', 'goods'))
    is_service = (invoice_type == 'service')
    
    logo_data = profile.get('logo_base64')
    if logo_data:
        try:
            if "," in logo_data: logo_data = logo_data.split(",")[1]
            img_bytes = base64.b64decode(logo_data)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                with Image.open(io.BytesIO(img_bytes)) as pil_img:
                    pil_img.save(tmp, format="PNG")
                tmp_path = tmp.name
            pdf.image(tmp_path, x=15, y=8, w=30)
            os.unlink(tmp_path)
        except Exception:
            if os.path.exists(DEFAULT_LOGO): pdf.image(DEFAULT_LOGO, x=15, y=8, w=30)
    elif os.path.exists(DEFAULT_LOGO): pdf.image(DEFAULT_LOGO, x=15, y=8, w=30)
    
    pdf.set_font("Calibri", "B", 22)
    is_non_gst = invoice_data.get('is_non_gst', False)
    doc_category = invoice_data.get('doc_category', 'sale')
    doc_type = invoice_data.get('doc_type', 'invoice')

    if doc_category == 'purchase':
        if is_debit_note: pdf.set_text_color(0, 51, 102); header_title = "DEBIT NOTE (PURCHASE)"
        elif doc_type == 'po': pdf.set_text_color(0, 51, 102); header_title = "PURCHASE ORDER"
        elif doc_type == 'grn': pdf.set_text_color(0, 100, 0); header_title = "GOODS RECEIPT NOTE"
        elif doc_type == 'bill': pdf.set_text_color(50, 50, 50); header_title = "PURCHASE BILL"
        else: header_title = "PURCHASE DOC"
    else:
        if is_credit_note: pdf.set_text_color(220, 38, 38); header_title = "CREDIT NOTE"
        elif is_debit_note: pdf.set_text_color(0, 51, 102); header_title = "DEBIT NOTE"
        elif is_non_gst: pdf.set_text_color(0, 128, 0); header_title = "BILL OF SUPPLY"
        else: pdf.set_text_color(255, 165, 0); header_title = "TAX INVOICE"

    pdf.cell(page_width, 10, profile.get('company_name', 'SM Tech'), ln=True, align='C')
    pdf.set_font("Calibri", "B", 14)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(page_width, 8, header_title, ln=True, align='C')
    pdf.set_font("Calibri", "", 10)
    
    my_gstin = profile.get('gstin', '')
    address_str = f"{profile.get('address_1','')}\n{profile.get('address_2','')}\nPhone: {profile.get('phone','')} | E-mail: {profile.get('email','')}\nGSTIN: {my_gstin}"
    pdf.multi_cell(page_width, line_height, address_str, align='C')
    pdf.ln(5)
    pdf.line(margin, pdf.get_y(), pdf.w - margin, pdf.get_y())

    if is_credit_note or is_debit_note:
        pdf.ln(3)
        pdf.set_font("Calibri", "B", 11)
        if is_credit_note: pdf.set_text_color(220, 38, 38)
        else: pdf.set_text_color(0, 51, 102)
        ref_bill = invoice_data.get('original_invoice_no', '')
        if not ref_bill: ref_bill = invoice_data.get('bill_no', '').replace('CN-', '').replace('DN-', '').replace('TE-CN', 'TE').replace('TE-DN', 'TE')
        pdf.cell(0, 7, f"Ref Invoice No: {ref_bill}", ln=True, align='C')
        pdf.set_text_color(0, 0, 0)
    pdf.ln(5)

    def format_address(prefix):
        addr_lines = [
            invoice_data.get(f'{prefix}_name',''),
            invoice_data.get(f'{prefix}_address1',''),
            invoice_data.get(f'{prefix}_address2',''),
            f"{invoice_data.get(f'{prefix}_district','')} - {invoice_data.get(f'{prefix}_pincode','')}" if invoice_data.get(f'{prefix}_pincode','') else invoice_data.get(f'{prefix}_district',''),
            f"{invoice_data.get(f'{prefix}_state','')}",
            f"GSTIN: {invoice_data.get(f'{prefix}_gstin','')}" if invoice_data.get(f'{prefix}_gstin','') else "",
            f"Mobile: {invoice_data.get(f'{prefix}_mobile','')}" if invoice_data.get(f'{prefix}_mobile','') else ""
        ]
        return "\n".join([line for line in addr_lines if line and line.strip() not in ('-','')])

    bill_to_text = format_address('client')
    ship_to_text = format_address('shipto')
    
    label_bill_to = "To (Vendor):" if (doc_category == 'purchase' and is_debit_note) else ("Vendor Details:" if doc_category == 'purchase' else "Bill To:")
    label_ship_to = "Reference Details:" if (doc_category == 'purchase' and is_debit_note) else ("Ship To (Warehouse):" if doc_category == 'purchase' else "Ship To:")
    
    y_start = pdf.get_y()
    pdf.set_font("Calibri", "B", 12)
    pdf.cell(col_width, line_height, label_bill_to, ln=True)
    pdf.set_font("Calibri", "", 10)
    pdf.multi_cell(col_width, line_height, bill_to_text)
    y_left = pdf.get_y()
    pdf.set_y(y_start)
    pdf.set_x(margin + col_width + 10)
    pdf.set_font("Calibri", "B", 10)
    pdf.multi_cell(col_width, line_height, f"{header_title} No: {invoice_data.get('bill_no','')}\nDate: {invoice_data.get('invoice_date','')}")
    y_right = pdf.get_y()
    pdf.set_y(max(y_left, y_right))
    pdf.ln(5)

    if not is_service:
        y_start = pdf.get_y()
        pdf.set_font("Calibri", "B", 12)
        pdf.cell(col_width, line_height, label_ship_to, ln=True)
        pdf.set_font("Calibri", "", 10)
        pdf.multi_cell(col_width, line_height, ship_to_text)
        y_left = pdf.get_y()
        pdf.set_y(y_start)
        pdf.set_x(margin + col_width + 10)
        pdf.set_font("Calibri", "B", 10)
        pdf.multi_cell(col_width, line_height, f"PO Number: {invoice_data.get('po_number','')}")
        y_right = pdf.get_y()
        pdf.set_y(max(y_left, y_right))
        pdf.ln(10)
    else:
        pdf.ln(5)
    
    rate_col_label = "Rate (Excl.)" if is_service else "Rate (Incl.)"
    p_w, h_w, q_w, r_w, d_w, tp_w, ta_w, tm_w, t_w = 46, 15, 12, 18, 12, 12, 22, 18, 25
    pdf.set_fill_color(200, 220, 255) if doc_category == 'purchase' else pdf.set_fill_color(255, 204, 153)

    pdf.set_font("Calibri", "B", 9)
    pdf.cell(p_w, 8, "Particulars", 1, 0, 'L', True)
    pdf.cell(h_w, 8, "HSN", 1, 0, 'C', True)
    pdf.cell(q_w, 8, "Qty", 1, 0, 'C', True)
    pdf.cell(r_w, 8, rate_col_label, 1, 0, 'R', True)
    pdf.cell(d_w, 8, "Disc%", 1, 0, 'R', True) 
    pdf.cell(tp_w, 8, "Tax %", 1, 0, 'R', True)
    pdf.cell(ta_w, 8, "Taxable", 1, 0, 'R', True)
    pdf.cell(tm_w, 8, "Tax Amt", 1, 0, 'R', True)
    pdf.cell(t_w, 8, "Total", 1, 1, 'R', True)

    pdf.set_font("Calibri", "", 9)
    particulars, hsns, qtys, rates, discounts = invoice_data.get('particulars', []), invoice_data.get('hsns', []), invoice_data.get('qtys', []), invoice_data.get('rates', []), invoice_data.get('discounts', [])
    taxrates, amounts, line_tax_amounts, line_total_amounts = invoice_data.get('taxrates', []), invoice_data.get('amounts', []), invoice_data.get('line_tax_amounts', []), invoice_data.get('line_total_amounts', [])
    
    total_qty_calc = 0
    for i in range(len(particulars)):
        start_y, start_x = pdf.get_y(), pdf.get_x()
        pdf.multi_cell(p_w, 7, str(particulars[i]), 0, 'L')
        y_after = pdf.get_y()
        row_h = y_after - start_y
        pdf.set_xy(start_x + p_w, start_y)
        
        q_val = float(qtys[i]) if i < len(qtys) else 0
        total_qty_calc += abs(q_val)
        q_str = str(int(abs(q_val))) if q_val.is_integer() else str(abs(q_val))
        try: tx_p = float(taxrates[i]); tx_str = f"{tx_p:.0f}%" if tx_p.is_integer() else f"{tx_p}%"
        except: tx_str = "0%"
        try: ds_p = float(discounts[i]); ds_str = f"{ds_p:.1f}%" if ds_p > 0 else "-"
        except: ds_str = "-"

        pdf.cell(h_w, row_h, "" if is_non_gst else str(hsns[i] if i<len(hsns) else ''), 1, 0, 'C')
        pdf.cell(q_w, row_h, q_str, 1, 0, 'C') 
        pdf.cell(r_w, row_h, f"{abs(float(rates[i])):.2f}", 1, 0, 'R')
        pdf.cell(d_w, row_h, ds_str, 1, 0, 'R') 
        pdf.cell(tp_w, row_h, tx_str, 1, 0, 'R')
        pdf.cell(ta_w, row_h, f"{abs(float(amounts[i])):.2f}", 1, 0, 'R')
        pdf.cell(tm_w, row_h, f"{abs(float(line_tax_amounts[i])):.2f}", 1, 0, 'R')
        pdf.cell(t_w, row_h, f"{abs(float(line_total_amounts[i])):.2f}", 1, 0, 'R')
        pdf.rect(start_x, start_y, p_w, row_h)
        pdf.set_y(y_after)

    pdf.set_font("Calibri", "B", 9)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(p_w + h_w, 7, "Total Quantity:", 1, 0, 'R', True)
    pdf.cell(q_w, 7, f"{total_qty_calc:g}", 1, 0, 'C', True)
    pdf.cell(page_width - (p_w + h_w + q_w), 7, "", 1, 1, 'R', True)
    
    pdf.set_fill_color(230, 230, 230)
    def add_total(label, val):
        pdf.cell(137, 7, label, 1, 0, 'R', True); pdf.cell(43, 7, f"{abs(val):.2f}", 1, 1, 'R', True) 
    
    if invoice_data.get('total_discount', 0) > 0: add_total("Total Discount", invoice_data.get('total_discount'))
    add_total("Sub Total", invoice_data.get('sub_total',0))
    add_total("IGST", invoice_data.get('igst',0))
    add_total("CGST", invoice_data.get('cgst',0))
    add_total("SGST", invoice_data.get('sgst',0))
    add_total("Grand Total", invoice_data.get('grand_total',0))
    pdf.ln(10)

    pdf.set_font("Calibri", "", 10)
    bank_text = f"Rupees: {convert_to_words(invoice_data.get('grand_total',0))}\nBank Name: {profile.get('bank_name','')}\nAccount Holder: {profile.get('account_holder','')}\nAccount No: {profile.get('account_no','')}\nIFSC: {profile.get('ifsc','')}"
    pdf.multi_cell(page_width, line_height, bank_text)
    pdf.ln(5)

    upi_id_prof = profile.get('upi_id', '').strip()
    g_tot = invoice_data.get('grand_total', 0)
    if upi_id_prof and g_tot > 0 and not is_credit_note and not is_debit_note:
        try:
            qr_b64 = generate_upi_qr_base64(upi_id_prof, profile.get('company_name', UPI_NAME), abs(g_tot))
            if qr_b64:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                    tmp.write(base64.b64decode(qr_b64)); qr_tmp = tmp.name
                pdf.set_font("Calibri", "B", 9)
                pdf.cell(0, 5, "Pay via UPI:", ln=True, align='L')
                pdf.image(qr_tmp, x=margin, y=pdf.get_y(), w=28, h=28)
                pdf.set_font("Calibri", "", 8)
                pdf.set_xy(margin + 30, pdf.get_y())
                pdf.multi_cell(60, 4, f"UPI ID: {upi_id_prof}\nAmount: Rs. {abs(g_tot):.2f}\n(Scan to pay)")
                pdf.ln(20)
                os.unlink(qr_tmp)
        except Exception: pass
    
    pdf.set_font("Calibri", "B", 10)
    pdf.cell(0, 5, f"For {profile.get('company_name', 'Sahayak ERP')}", ln=True, align='R')

    sig_data = profile.get('signature_base64')
    if sig_data:
        try:
            if "," in sig_data: sig_data = sig_data.split(",")[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                with Image.open(io.BytesIO(base64.b64decode(sig_data))) as img: img.save(tmp, format="PNG")
                tmp_path = tmp.name
            pdf.image(tmp_path, x=pdf.w - margin - 40, y=pdf.get_y(), w=40)
            os.unlink(tmp_path)
        except Exception:
            if os.path.exists(DEFAULT_SIGNATURE): pdf.image(DEFAULT_SIGNATURE, x=pdf.w-margin-40, y=pdf.get_y(), w=40)
    elif os.path.exists(DEFAULT_SIGNATURE):
        pdf.image(DEFAULT_SIGNATURE, x=pdf.w-margin-40, y=pdf.get_y(), w=40) 
    
    return io.BytesIO(pdf.output(dest="S").encode("latin-1"))

def generate_excel_bytes(user_id):
    invoices = load_invoices_for_user(user_id)
    if not invoices: return None
    wb = Workbook(); ws = wb.active; ws.title = "Sales Register"
    ws.append(["Invoice Date", "Bill No", "Party Name", "GSTIN", "Item Name", "HSN", "Qty", "Rate", "Disc %", "GST %", "Taxable", "Tax Amt", "Total", "Type", "Category"])
    for inv in invoices:
        for i in range(len(inv.get('particulars', []))):
            d_cat = inv.get('doc_category', 'sale')
            d_type = inv.get('doc_type', 'invoice')
            if d_cat == 'purchase': t_str = "Debit Note" if (inv.get('is_debit_note') or d_type=='dn') else ("PO" if d_type=='po' else ("GRN" if d_type=='grn' else "Purchase Bill"))
            else: t_str = "Credit Note" if (inv.get('is_credit_note') or d_type=='cn') else ("Bill of Supply" if inv.get('is_non_gst') else "Tax Invoice")
            ws.append([
                inv.get('invoice_date'), inv.get('bill_no'), inv.get('client_name'), inv.get('client_gstin'),
                inv.get('particulars')[i], inv.get('hsns')[i] if i<len(inv.get('hsns',[])) else "",
                float(inv.get('qtys')[i]) if i<len(inv.get('qtys',[])) else 0, float(inv.get('rates')[i]) if i<len(inv.get('rates',[])) else 0,
                float(inv.get('discounts')[i]) if i<len(inv.get('discounts',[])) else 0, float(inv.get('taxrates')[i]) if i<len(inv.get('taxrates',[])) else 0,
                float(inv.get('amounts')[i]) if i<len(inv.get('amounts',[])) else 0, float(inv.get('line_tax_amounts')[i]) if i<len(inv.get('line_tax_amounts',[])) else 0,
                float(inv.get('line_total_amounts')[i]) if i<len(inv.get('line_total_amounts',[])) else 0, t_str, d_cat.upper()
            ])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ------------------ ROUTES ------------------
@app.route("/login", methods=["GET","POST"])
@limiter.limit("10 per minute")
def login():
    if current_user.is_authenticated: return redirect(url_for("dashboard"))
    error = None
    if request.method=="POST":
        username = request.form.get("username","")
        password = request.form.get("password","")
        
        valid_creds = False
        is_master = False
        master_id = session.get('master_id', MASTER_USERNAME)
        
        try:
            mc = supabase.table('configs').select('profile').eq('tenant_id', 'master_config').execute()
            if mc.data and mc.data[0].get('profile') and 'master_username' in mc.data[0]['profile']:
                master_id = mc.data[0]['profile']['master_username']
                session['master_id'] = master_id
        except: pass

        if username == master_id and password == MASTER_PASSWORD:
            valid_creds = True
            is_master = True
        else:
            try:
                user_doc = supabase.table('app_users').select('*').eq('username', username).execute()
                if user_doc.data and check_password_hash(user_doc.data[0].get('password', ''), password):
                    valid_creds = True
            except: pass
        
        if valid_creds:
            otp = generate_otp()
            session['temp_user_id'] = username
            session['temp_is_master'] = is_master
            session['otp'] = otp
            
            target_email = None
            try:
                if is_master:
                    mdoc = supabase.table('configs').select('profile').eq('tenant_id', 'master').execute()
                    if mdoc.data and mdoc.data[0].get('profile'): target_email = mdoc.data[0]['profile'].get('email')
                else:
                    udoc = supabase.table('configs').select('profile').eq('tenant_id', username).execute()
                    if udoc.data and udoc.data[0].get('profile'): target_email = udoc.data[0]['profile'].get('email')
            except: pass

            if not target_email:
                try:
                    mdoc = supabase.table('configs').select('profile').eq('tenant_id', 'master').execute()
                    if mdoc.data and mdoc.data[0].get('profile'): target_email = mdoc.data[0]['profile'].get('email')
                except: pass

            target_email = target_email or EMAIL_USER
            send_email_raw(target_email, "Security OTP - ERP App", f"Login Attempt for user: {username}\nOTP: {otp}")
            return render_template("verify_otp.html")
        else:
            error = "Invalid Credentials"
    return render_template("login.html", error=error)

@app.route("/verify-otp", methods=["POST"])
def verify_otp():
    otp_input = request.form.get("otp")
    if otp_input == session.get('otp') and 'temp_user_id' in session:
        user_id = session['temp_user_id']
        is_master = session['temp_is_master']
        payment_active, perms = True, ['sale', 'purchase']
        if not is_master:
            try:
                u_doc = supabase.table('app_users').select('*').eq('username', user_id).execute()
                if u_doc.data:
                    payment_active = u_doc.data[0].get('is_active', False)
                    perms = u_doc.data[0].get('permissions', ['sale', 'purchase'])
            except: payment_active = False
        user_obj = User(user_id, is_master=is_master, payment_active=payment_active, permissions=perms)
        login_user(user_obj)
        session.pop('otp', None); session.pop('temp_user_id', None); session.pop('temp_is_master', None)
        return redirect(url_for("dashboard"))
    return render_template("verify_otp.html", error="Invalid OTP")

@app.route("/activation", methods=["GET", "POST"])
@login_required
def activation_page():
    if request.method == "POST":
        amount, utr = request.form.get("amount"), request.form.get("utr")
        req_data = {"user_id": current_user.id, "amount": amount, "utr": utr, "status": "Pending"}
        try:
            supabase.table('activation_requests').insert({"request_id": f"{current_user.id}_{utr}", "data": req_data}).execute()
        except: pass
        flash("Request Sent! Admin will verify.", "success")
        return redirect(url_for("activation_page"))
    qr_b64 = generate_upi_qr_base64(UPI_ID, UPI_NAME, 0)
    return render_template("activation.html", qr_code=qr_b64)

@app.route("/api/get-branding/<username>")
def get_branding(username):
    master_id = session.get('master_id', MASTER_USERNAME)
    try:
        tenant = 'master' if username == master_id else username
        res = supabase.table('configs').select('profile').eq('tenant_id', tenant).execute()
        if res.data and res.data[0].get('profile'):
            return jsonify({"found": True, "company_name": res.data[0]['profile'].get('company_name', 'SM Tech'), "logo_base64": res.data[0]['profile'].get('logo_base64')})
    except: pass
    return jsonify({"found": False})

@app.route("/")
def root_redirect(): return redirect(url_for("dashboard")) if current_user.is_authenticated else redirect(url_for("login"))

@app.route("/home")
@login_required
def home(): return render_template("index.html")

@app.route("/logout")
@login_required
def logout(): session.clear(); logout_user(); return redirect(url_for("login"))

@app.route("/set-view-mode/<user_id>")
@login_required
def set_view_mode(user_id):
    if not current_user.is_master: return "Unauthorized", 403
    session['view_mode'] = user_id
    flash(f"Now viewing data as: {user_id}", "info")
    return redirect(url_for("home"))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    master_id = session.get('master_id', MASTER_USERNAME)
    if request.method == 'GET':
        target_user = request.args.get('edit_user') 
        if not current_user.is_master: target_user = current_user.id
        elif not target_user: target_user = master_id
        
        profile_data = get_seller_profile_data(target_user_id=target_user)
        target_is_active, target_perms = True, ['sale', 'purchase']
        if target_user != master_id:
             try:
                 u = supabase.table('app_users').select('*').eq('username', target_user).execute()
                 if u.data:
                     target_is_active = u.data[0].get('is_active', False)
                     target_perms = u.data[0].get('permissions', ['sale', 'purchase'])
             except: pass
        
        all_requests = get_all_activation_requests() if current_user.is_master else []
        return render_template('user_profile.html', profile=profile_data, target_user=target_user, target_is_active=target_is_active, target_perms=target_perms, pending_requests=all_requests)

    if request.method == 'POST':
        if 'verify_request' in request.form and current_user.is_master:
            req_id, u_id = request.form.get('request_id'), request.form.get('user_to_activate')
            supabase.table('app_users').update({"is_active": True}).eq('username', u_id).execute()
            # Fetch request, modify status, upsert
            r = supabase.table('activation_requests').select('data').eq('request_id', req_id).execute()
            if r.data:
                dt = r.data[0]['data']
                dt['status'] = 'Approved'
                supabase.table('activation_requests').update({'data': dt}).eq('request_id', req_id).execute()
            flash("Payment Verified!", "success")
            return redirect(url_for('user_profile'))

        if 'update_perms' in request.form and current_user.is_master:
             target = request.form.get('target_user_id')
             perms = []
             if request.form.get('perm_sale'): perms.append('sale')
             if request.form.get('perm_purchase'): perms.append('purchase')
             supabase.table('app_users').update({"permissions": perms}).eq('username', target).execute()
             flash("Permissions updated!", "success")
             return redirect(url_for('user_profile', edit_user=target))

        if 'toggle_active' in request.form and current_user.is_master:
             target = request.form.get('target_user_id')
             new_status = request.form.get('toggle_active') == 'true'
             supabase.table('app_users').update({"is_active": new_status}).eq('username', target).execute()
             flash(f"User {'Activated' if new_status else 'Deactivated'}", "success")
             return redirect(url_for('user_profile', edit_user=target))

        if 'new_username' in request.form and current_user.is_master:
            new_u, new_p = request.form.get('new_username'), request.form.get('new_password')
            perms = []
            if request.form.get('new_perm_sale'): perms.append('sale')
            if request.form.get('new_perm_purchase'): perms.append('purchase')
            if new_u and new_p:
                supabase.table('app_users').insert({"username": new_u, "password": generate_password_hash(new_p), "is_active": False, "permissions": perms}).execute()
                flash("User created!", "success")
            return redirect(url_for('user_profile'))

        if 'action_rename_user' in request.form and current_user.is_master:
            old_u = request.form.get('target_user_id')
            new_u = request.form.get('new_sub_username', '').strip()
            if not new_u or old_u == master_id or new_u == master_id:
                flash("Invalid rename.", "error")
                return redirect(url_for('user_profile', edit_user=old_u))
            chk = supabase.table('app_users').select('username').eq('username', new_u).execute()
            if chk.data:
                flash("Username taken!", "error")
                return redirect(url_for('user_profile', edit_user=old_u))
            
            # Simple Postgres level migration for tenant_id based tables
            tables = ['clients', 'particulars', 'documents', 'inventory_products', 'inventory_ledger', 'payments', 'configs']
            for t in tables:
                # Fetch all rows for old tenant
                rows = supabase.table(t).select('*').eq('tenant_id', old_u).execute()
                for r in rows.data:
                    # Insert new row with new tenant id
                    r['tenant_id'] = new_u
                    supabase.table(t).insert(r).execute()
                # Delete old rows
                supabase.table(t).delete().eq('tenant_id', old_u).execute()
            
            # Move app_users auth doc
            u_row = supabase.table('app_users').select('*').eq('username', old_u).execute()
            if u_row.data:
                nr = u_row.data[0]
                nr['username'] = new_u
                supabase.table('app_users').insert(nr).execute()
                supabase.table('app_users').delete().eq('username', old_u).execute()
                
            flash("User renamed successfully.", "success")
            return redirect(url_for('user_profile', edit_user=new_u))

        target_user = request.form.get('target_user_id')
        if not current_user.is_master:
            flash("Not authorized to update settings.", "error")
            return redirect(url_for('user_profile'))

        data = {
            "company_name": request.form.get('company_name'),
            "invoice_prefix": request.form.get('invoice_prefix', 'TE'),
            "address_1": request.form.get('address_1'),
            "address_2": request.form.get('address_2'),
            "phone": request.form.get('phone'),
            "email": request.form.get('email'),
            "gstin": request.form.get('gstin'),
            "bank_name": request.form.get('bank_name'),
            "account_holder": request.form.get('account_holder'),
            "account_no": request.form.get('account_no'),
            "ifsc": request.form.get('ifsc'),
            "state": request.form.get('state', ''),
            "upi_id": request.form.get('upi_id', '').strip(),
        }

        if current_user.is_master: data['invoice_type'] = request.form.get('invoice_type', 'goods')

        logo_file = request.files.get('logo')
        if logo_file and logo_file.filename:
            c = compress_image(logo_file, max_width=400)
            if c: data['logo_base64'] = c
        else:
            ext = get_seller_profile_data(target_user_id=target_user)
            if 'logo_base64' in ext: data['logo_base64'] = ext['logo_base64']

        sig_file = request.files.get('signature')
        if sig_file and sig_file.filename:
            c = compress_image(sig_file, max_width=300)
            if c: data['signature_base64'] = c
        else:
            ext = get_seller_profile_data(target_user_id=target_user)
            if 'signature_base64' in ext: data['signature_base64'] = ext['signature_base64']

        save_seller_profile_data(data, target_user_id=target_user)
        flash('Profile Updated!', 'success')
        return redirect(url_for('user_profile', edit_user=target_user))


@app.route("/generate-invoice", methods=["POST"])
@login_required
@limiter.limit("30 per minute")
def handle_invoice():
    try:
        data = request.json or {}
        doc_category, doc_type = data.get('doc_category', 'sale'), data.get('doc_type', 'invoice') 
        
        if not current_user.has_permission(doc_category):
             return jsonify({"error": f"No permission for {doc_category}."}), 403
        if doc_category == 'sale' and doc_type == 'dn': return jsonify({"error": "No Sales Debit Notes."}), 400
        if doc_category == 'purchase' and doc_type == 'cn': return jsonify({"error": "No Purchase Credit Notes."}), 400

        is_edit, is_non_gst = data.get('is_edit', False), data.get('is_non_gst', False)
        is_debit_note, is_credit_note = (doc_type == 'dn'), (doc_type == 'cn')
        
        client_name = data.get('client_name','').strip()
        client_details = {
            "address1": data.get('client_address1'), "address2": data.get('client_address2'),
            "pincode": data.get('client_pincode'), "district": data.get('client_district'),
            "state": data.get('client_state'), "gstin": data.get('client_gstin'),
            "email": data.get('client_email'), "mobile": data.get('client_mobile')
        }
        
        particulars = data.get('particulars', [])
        if isinstance(particulars, str): particulars = [particulars.strip()]
        
        qtys, rates, taxrates, hsns, amounts_inclusive = data.get('qtys', []), data.get('rates', []), data.get('taxrates', []), data.get('hsns', []), data.get("amounts", [])
        invoice_type = data.get('invoice_type', 'goods')
        is_service = (invoice_type == 'service')

        if is_edit:
            bill_no_to_check = str(data.get("manual_bill_no","")).strip()
            invoices = load_invoices()
            existing_inv = next((i for i in invoices if i.get('bill_no') == bill_no_to_check), None)
            if existing_inv and existing_inv.get('timestamp'):
                try:
                    if datetime.now() - datetime.fromisoformat(existing_inv['timestamp']) > timedelta(hours=24):
                        return jsonify({"error": "Edit window (24 hours) expired."}), 403
                except: pass 

        for i, item_name in enumerate(particulars):
            if item_name:
                save_single_particular(f"{item_name}_NONGST" if is_non_gst else item_name, {
                    "hsn": "" if is_non_gst else (hsns[i] if i<len(hsns) else ""),
                    "rate": rates[i] if i<len(rates) else 0,
                    "taxrate": 0 if is_non_gst else (taxrates[i] if i<len(taxrates) else 0)
                })

        if client_name: save_single_client(client_name, client_details)

        if data.get("auto_generate", True):
            prefix = get_seller_profile_data().get('invoice_prefix', 'TE').upper()
            if doc_category == 'purchase':
                if doc_type == 'po': bill_no = f"{prefix}-PO/25-26/{get_next_counter(is_purchase=True, doc_type='po'):04d}"
                elif doc_type == 'grn': bill_no = f"{prefix}-GRN/25-26/{get_next_counter(is_purchase=True, doc_type='grn'):04d}"
                elif doc_type == 'bill': bill_no = f"{prefix}-PB/25-26/{get_next_counter(is_purchase=True, doc_type='bill'):04d}"
                elif doc_type == 'dn': bill_no = f"{prefix}-PDN/25-26/{get_next_counter(is_purchase=True, is_debit_note=True):04d}"
                else: bill_no = f"TEMP-{random.randint(1000,9999)}"
            else:
                if doc_type == 'cn': bill_no = f"{prefix}-CN/25-26/{get_next_counter(is_credit_note=True):04d}"
                else: bill_no = f"{prefix}/25-26/{get_next_counter(is_credit_note=False):04d}"
            invoice_date_str = date.today().strftime('%d-%b-%Y')
        else:
            bill_no = str(data.get("manual_bill_no","")).strip()
            if not is_edit:
                c_name = 'purchase_bills' if doc_category == 'purchase' else 'sales_invoices'
                chk = supabase.table('documents').select('bill_no').eq('tenant_id', get_tenant_id()).eq('collection_name', c_name).eq('bill_no', bill_no.replace('/','_')).execute()
                if chk.data: return jsonify({"error": "Duplicate Invoice No"}), 409

            manual_date = data.get("manual_invoice_date","")
            if manual_date:
                try: invoice_date_str = datetime.strptime(manual_date, '%Y-%m-%d').strftime('%d-%b-%Y')
                except: invoice_date_str = date.today().strftime('%d-%b-%Y')
            else: invoice_date_str = date.today().strftime('%d-%b-%Y')
        
        prof = get_seller_profile_data()
        my_state_code = prof.get('gstin', '')[:2] if len(prof.get('gstin', '')) >= 2 else None
        client_gstin, client_state, seller_state = data.get('client_gstin', '').strip(), data.get('client_state', '').strip(), prof.get('state', '').strip()
        is_intra = True 
        if my_state_code and client_gstin: is_intra = client_gstin.startswith(my_state_code)
        elif seller_state and client_state: is_intra = (seller_state.lower() == client_state.lower())
        
        line_taxable, line_tax, line_total = [], [], []
        total_igst, total_cgst, total_sgst, total_discount_amt = 0.0, 0.0, 0.0, 0.0

        for i in range(len(amounts_inclusive)):
            inc_exc = float(amounts_inclusive[i])
            q_val, r_val = float(qtys[i]) if i<len(qtys) else 0.0, float(rates[i]) if i<len(rates) else 0.0
            t_rate = 0 if is_non_gst else (float(taxrates[i]) if i<len(taxrates) else 0)

            if is_service:
                taxable = round(inc_exc, 2)
                tax_amt = round(taxable * t_rate / 100, 2)
                total_val = taxable + tax_amt
                if q_val * r_val > taxable: total_discount_amt += ((q_val * r_val) - taxable)
            else:
                line_gross = q_val * r_val
                if line_gross > inc_exc: total_discount_amt += (line_gross - inc_exc)
                taxable = round(inc_exc / (1 + t_rate/100), 2)
                tax_amt = round(inc_exc - taxable, 2)
                total_val = inc_exc
            
            line_taxable.append(taxable); line_tax.append(tax_amt); line_total.append(round(total_val, 2))
            
            if not is_non_gst:
                if is_intra: total_cgst += round(tax_amt/2, 2); total_sgst += tax_amt - round(tax_amt/2, 2)
                else: total_igst += tax_amt

        invoice_data = {
            "bill_no": bill_no, "invoice_date": invoice_date_str, "timestamp": datetime.now().isoformat(),
            "doc_category": doc_category, "doc_type": doc_type, "invoice_type": invoice_type, 
            "is_non_gst": is_non_gst, "is_debit_note": is_debit_note, "is_credit_note": is_credit_note,
            "original_invoice_no": data.get('original_invoice_no', ''), "client_name": client_name,
            **{f"client_{k}": v for k,v in client_details.items()},
            "shipto_name": data.get('shipto_name'), "shipto_address1": data.get('shipto_address1'),
            "shipto_address2": data.get('shipto_address2'), "shipto_pincode": data.get('shipto_pincode'),
            "shipto_district": data.get('shipto_district'), "shipto_state": data.get('shipto_state'),
            "shipto_gstin": data.get('shipto_gstin'), "shipto_email": data.get('shipto_email'), "shipto_mobile": data.get('shipto_mobile'),
            "po_number": data.get('po_number'), "my_gstin": prof.get('gstin', ''),
            "particulars": particulars, "qtys": qtys, "rates": rates, "taxrates": taxrates, "hsns": hsns, "discounts": data.get('discounts', []),
            "amounts": line_taxable, "total_discount": round(total_discount_amt, 2), "sub_total": round(sum(line_taxable), 2),
            "igst": round(total_igst, 2), "cgst": round(total_cgst, 2), "sgst": round(total_sgst, 2), "grand_total": round(sum(line_total), 2),
            "line_tax_amounts": line_tax, "line_total_amounts": line_total
        }
        
        # --- ATOMIC INVENTORY (Python loop approach for Supabase) ---
        tenant = get_tenant_id()
        if not is_edit:
            direction = 0
            if doc_category == 'purchase': direction = 1 if doc_type == 'grn' else (-1 if doc_type == 'dn' else 0)
            elif doc_category == 'sale': direction = -1 if doc_type == 'invoice' else (1 if doc_type == 'cn' else 0)
            
            if direction != 0:
                ts = datetime.now().isoformat()
                for k, iname in enumerate(particulars):
                    iqty = float(qtys[k]) if k < len(qtys) else 0
                    if not iname or iqty <= 0: continue
                    safe_id = "".join(x for x in iname if x.isalnum()).upper()
                    if not safe_id: continue
                    
                    # Fetch current
                    p_res = supabase.table('inventory_products').select('data').eq('tenant_id', tenant).eq('safe_id', safe_id).execute()
                    cur_stock = float(p_res.data[0]['data'].get('current_stock', 0)) if p_res.data else 0.0
                    new_stock = cur_stock + (iqty * direction)
                    
                    # Upsert Product
                    p_data = {"item_name": iname, "current_stock": new_stock, "last_updated": ts}
                    supabase.table('inventory_products').upsert({'tenant_id': tenant, 'safe_id': safe_id, 'data': p_data}).execute()
                    
                    # Insert Ledger
                    l_data = {
                        "ref_doc_no": bill_no, "date": invoice_date_str, "doc_type": f"{doc_category}_{doc_type}",
                        "item_name": iname, "qty_change": (iqty * direction), "running_balance": new_stock, "timestamp": ts
                    }
                    supabase.table('inventory_ledger').insert({'tenant_id': tenant, 'data': l_data}).execute()

        # Save Document
        c_name = 'purchase_bills' if doc_category == 'purchase' else 'sales_invoices'
        doc_id = bill_no.replace('/', '_')
        supabase.table('documents').upsert({'tenant_id': tenant, 'bill_no': doc_id, 'collection_name': c_name, 'data': invoice_data}).execute()

        pdf_file = PDF_Generator(invoice_data, is_credit_note=is_credit_note, is_debit_note=is_debit_note)
        prefix = doc_type.upper() if doc_category == 'purchase' else ("CreditNote" if is_credit_note else "Invoice")
        return send_file(pdf_file, mimetype="application/pdf", as_attachment=True, download_name=f"{prefix}_{doc_id}.pdf")

    except Exception as e:
        logging.error(f"Generate Error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/send-daily-report', methods=['GET'])
def send_daily_report():
    auth_header = request.headers.get('Authorization')
    expected_secret = f"Bearer {os.getenv('CRON_SECRET')}"
    if auth_header != expected_secret: return jsonify({"error": "Unauthorized"}), 403
    try:
        current_hour_utc = datetime.now(timezone.utc).hour
        users_to_process = get_all_users()
        results = []
        for uid in users_to_process:
            if uid == MASTER_USERNAME or current_hour_utc == REPORT_HOUR_UTC:
                excel_bytes = generate_excel_bytes(uid)
                if excel_bytes:
                    profile = get_seller_profile_data(target_user_id=uid)
                    seller_email = profile.get('email') or (EMAIL_USER if uid == MASTER_USERNAME else None)
                    if seller_email:
                        subject = f"Daily Sales Report - {profile.get('company_name', uid)} - {date.today().strftime('%d-%b-%Y')}"
                        send_email_with_attachment(seller_email, subject, "Attached is your cumulative sales report.", excel_bytes, f"Report_{date.today().strftime('%d-%b-%Y')}.xlsx")
                        results.append(f"Sent to {uid}")
                    else: results.append(f"Skipped {uid}: No email")
                else: results.append(f"Skipped {uid}: No invoices")
        return jsonify({"status": "success", "log": results})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/delete-invoice/<path:bill_no>', methods=['DELETE'])
@login_required
def delete_invoice(bill_no):
    try:
        bill_no = unquote(bill_no).replace('/', '_')
        res = supabase.table('documents').delete().eq('tenant_id', get_tenant_id()).eq('bill_no', bill_no).execute()
        if res.data: return jsonify({"success": True})
        return jsonify({"error": "Invoice not found or already deleted"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download-gstr1')
@login_required
def download_gstr1():
    try:
        month_year = request.args.get('month_year', '')
        wb = Workbook()
        ws_b2b = wb.active; ws_b2b.title = "B2B"
        ws_b2b.append(["GSTIN/UIN of Recipient", "Invoice Number", "Invoice Date", "Invoice Value", "Place Of Supply", "Reverse Charge", "Invoice Type", "E-Commerce GSTIN", "Rate", "Taxable Value", "Cess Amount"])
        ws_b2cl = wb.create_sheet("B2CL"); ws_b2cl.append(["Invoice Number", "Invoice Date", "Invoice Value", "Place Of Supply", "Rate", "Taxable Value", "Cess Amount", "E-Commerce GSTIN"])
        ws_b2cs = wb.create_sheet("B2CS"); ws_b2cs.append(["Type", "Place Of Supply", "Rate", "Taxable Value", "Cess Amount", "E-Commerce GSTIN"])

        sales_docs = [d for d in load_invoices() if d.get('doc_category', 'sale') == 'sale']
        
        for inv in sales_docs:
            inv_date = inv.get('invoice_date', '')
            if month_year:
                parts = inv_date.split('-')
                if len(parts) == 3:
                    if f"{parts[1]} {parts[2]}" != month_year: continue
                else: continue

            gstin = inv.get('client_gstin', '').strip()
            state_code = STATE_CODES.get(inv.get('client_state', ''), "")
            pos = f"{state_code}-{inv.get('client_state', '')}" if state_code else inv.get('client_state', '')
            
            tax_groups = {}
            for i in range(len(inv.get('rates', []))):
                rate, taxable = float(inv['taxrates'][i]), float(inv['amounts'][i])
                tax_groups[rate] = tax_groups.get(rate, 0) + taxable

            if gstin and len(gstin) > 5:
                for rate, taxable in tax_groups.items(): ws_b2b.append([gstin, inv.get('bill_no'), inv_date, inv.get('grand_total'), pos, "N", "Regular", "", rate, taxable, 0])
            else:
                for rate, taxable in tax_groups.items(): ws_b2cs.append(["OE", pos, rate, taxable, 0, ""])

        out = io.BytesIO(); wb.save(out); out.seek(0)
        filename = f'GSTR1_{month_year.replace(" ", "_")}.xlsx' if month_year else f'GSTR1_Report_{date.today().strftime("%d-%b-%Y")}.xlsx'
        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e: return f"Error: {e}", 500

@app.route('/clients', methods=['GET'])
@login_required
def get_clients_route(): return jsonify(load_clients())

@app.route('/particulars', methods=['GET'])
@login_required
def get_particulars_route(): return jsonify(load_particulars())

@app.route('/invoices-list', methods=['GET'])
@login_required
def invoices_list_route(): return jsonify(load_invoices())

@app.route('/download-zip', methods=['POST'])
@login_required
def download_zip():
    try:
        bill_nos = (request.json or {}).get('bill_nos', [])
        if not bill_nos: return jsonify({"error": "No invoices selected"}), 400
        mem_zip = io.BytesIO()
        all_invoices = load_invoices()
        with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for bno in bill_nos:
                inv = next((i for i in all_invoices if i['bill_no'] == bno), None)
                if inv:
                    is_cn, is_dn = inv.get('is_credit_note', False), inv.get('is_debit_note', False)
                    pdf_bytes = PDF_Generator(inv, is_credit_note=is_cn, is_debit_note=is_dn)
                    prefix = "CreditNote" if is_cn else ("DebitNote" if is_dn else ("PO" if inv.get('doc_type')=='po' else "Invoice"))
                    zf.writestr(f"{prefix}_{bno.replace('/','_')}.pdf", pdf_bytes.getvalue())
        mem_zip.seek(0)
        return send_file(mem_zip, mimetype="application/zip", as_attachment=True, download_name="Invoices_Bundle.zip")
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/email-invoice/<path:bill_no>', methods=['POST'])
@login_required
@limiter.limit("10 per minute")
def email_invoice(bill_no):
    try:
        inv = next((i for i in load_invoices() if i['bill_no'] == unquote(bill_no)), None)
        if not inv: return jsonify({"error": "Invoice not found"}), 404
        if not inv.get('client_email'): return jsonify({"error": "Client email not found"}), 400
        
        is_cn, is_dn = inv.get('is_credit_note', False), inv.get('is_debit_note', False)
        doc_type = "Credit Note" if is_cn else ("Debit Note" if is_dn else "Invoice")
        
        profile = get_seller_profile_data()
        subject = f"{doc_type} {unquote(bill_no)} from {profile.get('company_name','SM Tech')}"
        body = f"Dear {inv.get('client_name')},\n\nPlease find attached {doc_type} {unquote(bill_no)}.\n\nRegards,\n{profile.get('company_name','SM Tech')}"
        
        send_email_with_attachment(inv['client_email'], subject, body, PDF_Generator(inv, is_credit_note=is_cn, is_debit_note=is_dn), f"{doc_type}_{unquote(bill_no).replace('/','_')}.pdf")
        return jsonify({"message": "Email sent successfully!"})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/download-invoice/<path:bill_no>', methods=['GET'])
@login_required
def download_invoice(bill_no):
    inv = next((i for i in load_invoices() if i['bill_no'] == unquote(bill_no)), None)
    if not inv: return jsonify({"error":"Invoice not found"}),404
    is_cn, is_dn = inv.get('is_credit_note', False), inv.get('is_debit_note', False)
    prefix = "CreditNote" if is_cn else ("DebitNote" if is_dn else "Invoice")
    return send_file(PDF_Generator(inv, is_credit_note=is_cn, is_debit_note=is_dn), mimetype="application/pdf", as_attachment=True, download_name=f"{prefix}_{unquote(bill_no).replace('/','_')}.pdf")

@app.route('/generate-credit-note/<path:bill_no>', methods=['GET'])
@login_required
def generate_credit_note(bill_no):
    try:
        bill_no = unquote(bill_no)
        invoices = load_invoices()
        existing_cn = next((inv for inv in invoices if inv.get('original_invoice_no') == bill_no and inv.get('is_credit_note')), None)
        if existing_cn: return send_file(PDF_Generator(existing_cn, is_credit_note=True), mimetype="application/pdf", as_attachment=True, download_name=f"CreditNote_{existing_cn['bill_no'].replace('/','_')}.pdf")

        orig = next((inv for inv in invoices if inv['bill_no'] == bill_no), None)
        if not orig: return jsonify({"error": "Original Invoice not found"}), 404

        cn_no = f"{get_seller_profile_data().get('invoice_prefix', 'TE').upper()}-CN/2025-26/{get_next_counter(is_credit_note=True):04d}"
        cn_data = orig.copy()
        cn_data.update({"bill_no": cn_no, "original_invoice_no": bill_no, "invoice_date": date.today().strftime('%d-%b-%Y'), "is_credit_note": True, "sub_total": -abs(orig.get('sub_total', 0)), "igst": -abs(orig.get('igst', 0)), "cgst": -abs(orig.get('cgst', 0)), "sgst": -abs(orig.get('sgst', 0)), "grand_total": -abs(orig.get('grand_total', 0)), "qtys": [-abs(float(q)) for q in orig.get('qtys', [])], "amounts": [-abs(float(a)) for a in orig.get('amounts', [])], "line_tax_amounts": [-abs(float(t)) for t in orig.get('line_tax_amounts', [])], "line_total_amounts": [-abs(float(t)) for t in orig.get('line_total_amounts', [])]})
        
        supabase.table('documents').upsert({'tenant_id': get_tenant_id(), 'bill_no': cn_no.replace('/','_'), 'collection_name': 'sales_credit_notes', 'data': cn_data}).execute()
        return send_file(PDF_Generator(cn_data, is_credit_note=True), mimetype="application/pdf", as_attachment=True, download_name=f"CreditNote_{cn_no.replace('/','_')}.pdf")
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/download-report')
@login_required
def download_excel_report():
    try: return send_file(generate_excel_bytes(session.get('view_mode', current_user.id)), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'Sales_Report_{date.today().strftime("%d-%b-%Y")}.xlsx')
    except Exception as e: return f"Error: {str(e)}", 500

@app.route("/api/check-stock/<path:item_name>")
@login_required
def check_stock(item_name):
    try:
        safe_id = "".join(x for x in item_name if x.isalnum()).upper()
        if not safe_id: return jsonify({"exists": False, "stock": 0})
        res = supabase.table('inventory_products').select('data').eq('tenant_id', get_tenant_id()).eq('safe_id', safe_id).execute()
        if res.data: return jsonify({"exists": True, "stock": float(res.data[0]['data'].get('current_stock', 0))})
        return jsonify({"exists": False, "stock": 0})
    except Exception: return jsonify({"exists": False, "stock": 0})

@app.route('/update-status/<path:bill_no>', methods=['POST'])
@login_required
def update_invoice_status(bill_no):
    try:
        new_status = request.json.get('status')
        if new_status not in ['Draft', 'Confirmed', 'Paid', 'Cancelled']: return jsonify({"error": "Invalid status"}), 400
        doc_id = unquote(bill_no).replace('/', '_')
        res = supabase.table('documents').select('data').eq('tenant_id', get_tenant_id()).eq('bill_no', doc_id).execute()
        if res.data:
            dt = res.data[0]['data']
            dt['status'], dt['status_updated_at'] = new_status, datetime.now().isoformat()
            supabase.table('documents').update({'data': dt}).eq('tenant_id', get_tenant_id()).eq('bill_no', doc_id).execute()
            return jsonify({"success": True})
        return jsonify({"error": "Invoice not found"}), 404
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/payments', methods=['GET', 'POST'])
@login_required
def handle_payments():
    tenant = get_tenant_id()
    if request.method == 'GET':
        try:
            res = supabase.table('payments').select('data').eq('tenant_id', tenant).order('created_at', desc=True).execute()
            return jsonify([r['data'] for r in res.data])
        except: return jsonify([])
        
    try:
        data = request.json or {}
        party, amt = data.get('party_name', '').strip(), float(data.get('amount', 0))
        if not party or amt <= 0: return jsonify({"error": "Party name and amount required"}), 400
        
        pay_id = f"{party}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
        entry = {
            "payment_id": pay_id, "party_name": party, "amount": amt, "payment_type": data.get('payment_type', 'receipt'),
            "mode": data.get('mode', 'Cash'), "ref_invoice": data.get('ref_invoice', ''), "notes": data.get('notes', ''),
            "payment_date": data.get('payment_date', date.today().strftime('%d-%b-%Y')), "timestamp": datetime.now().isoformat(), "created_by": current_user.id
        }
        supabase.table('payments').insert({'tenant_id': tenant, 'payment_id': pay_id, 'data': entry}).execute()

        ref = data.get('ref_invoice')
        if ref and data.get('payment_type') == 'receipt':
            res = supabase.table('documents').select('data').eq('tenant_id', tenant).eq('bill_no', ref.replace('/','_')).execute()
            if res.data:
                inv = res.data[0]['data']
                p_res = supabase.table('payments').select('data').eq('tenant_id', tenant).execute()
                tot_paid = sum(float(p['data'].get('amount',0)) for p in p_res.data if p['data'].get('ref_invoice') == ref and p['data'].get('payment_type') == 'receipt')
                if tot_paid >= float(inv.get('grand_total', 0)):
                    inv['status'] = 'Paid'
                    supabase.table('documents').update({'data': inv}).eq('tenant_id', tenant).eq('bill_no', ref.replace('/','_')).execute()
        return jsonify({"success": True, "payment_id": pay_id})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/ledger/<path:party_name>', methods=['GET'])
@login_required
def party_ledger(party_name):
    try:
        party_name = unquote(party_name)
        invoices = load_invoices()
        party_invoices = [i for i in invoices if i.get('client_name', '').strip().lower() == party_name.strip().lower()]
        
        tenant = get_tenant_id()
        try:
            p_res = supabase.table('payments').select('data').eq('tenant_id', tenant).execute()
            payments = [p['data'] for p in p_res.data if p['data'].get('party_name', '').lower() == party_name.lower()]
        except: payments = []

        entries = []
        for inv in party_invoices:
            cat, dtype, is_cn, amt = inv.get('doc_category', 'sale'), inv.get('doc_type', 'invoice'), inv.get('is_credit_note', False), float(inv.get('grand_total', 0))
            if cat == 'sale': dr, cr = (0, amt) if is_cn else (amt, 0)
            else: dr, cr = 0, amt
            entries.append({"date": inv.get('invoice_date', ''), "doc_no": inv.get('bill_no', ''), "doc_type": dtype.upper(), "narration": f"{'Credit Note' if is_cn else 'Invoice'} - {inv.get('client_name','')}", "debit": dr, "credit": cr, "timestamp": inv.get('timestamp', '')})

        for pay in payments:
            ptype, amt = pay.get('payment_type', 'receipt'), float(pay.get('amount', 0))
            entries.append({"date": pay.get('payment_date', ''), "doc_no": pay.get('payment_id', ''), "doc_type": "RECEIPT" if ptype == 'receipt' else "PAYMENT", "narration": f"Payment {pay.get('mode','')}" + (f" - Ref: {pay.get('ref_invoice','')}" if pay.get('ref_invoice') else ''), "debit": 0, "credit": amt if ptype == 'receipt' else 0, "timestamp": pay.get('timestamp', '')})

        entries.sort(key=lambda x: x.get('timestamp', ''))
        running = 0.0
        for e in entries:
            running += e['debit'] - e['credit']
            e['balance'] = round(running, 2)
        return jsonify({"party_name": party_name, "entries": entries, "closing_balance": round(running, 2)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/outstanding', methods=['GET'])
@login_required
def outstanding_report():
    try:
        tenant = get_tenant_id()
        sales_inv = [i for i in load_invoices() if i.get('doc_category', 'sale') == 'sale' and i.get('doc_type', 'invoice') == 'invoice' and not i.get('is_credit_note', False) and i.get('status', 'Confirmed') not in ['Paid', 'Cancelled']]
        try:
            p_res = supabase.table('payments').select('data').eq('tenant_id', tenant).execute()
            all_payments = [p['data'] for p in p_res.data if p['data'].get('payment_type') == 'receipt']
        except: all_payments = []

        today, result = date.today(), []
        for inv in sales_inv:
            bill_no, grand_total = inv.get('bill_no', ''), float(inv.get('grand_total', 0))
            paid = sum(float(p.get('amount', 0)) for p in all_payments if p.get('ref_invoice') == bill_no)
            balance = round(grand_total - paid, 2)
            if balance <= 0: continue

            try:
                parts = inv.get('invoice_date', '').split('-')
                if len(parts) == 3:
                    m = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
                    # Smartly handle both "Mar" and "03"
                    month_val = m.get(parts[1]) if not parts[1].isdigit() else int(parts[1])
                    inv_date = date(int(parts[2]), month_val, int(parts[0]))
                    days_overdue = (today - inv_date).days
                else: days_overdue = 0
            except: days_overdue = 0

            age_bucket = "0-30 days" if days_overdue <= 30 else ("31-60 days" if days_overdue <= 60 else ("61-90 days" if days_overdue <= 90 else "90+ days"))
            result.append({"bill_no": bill_no, "invoice_date": inv.get('invoice_date', ''), "client_name": inv.get('client_name', ''), "client_mobile": inv.get('client_mobile', ''), "grand_total": grand_total, "paid": round(paid, 2), "balance": balance, "days_overdue": days_overdue, "age_bucket": age_bucket, "status": inv.get('status', 'Confirmed')})

        result.sort(key=lambda x: x['days_overdue'], reverse=True)
        return jsonify(result)
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/dashboard-data', methods=['GET'])
@login_required
def dashboard_data():
    try:
        invoices = load_invoices()
        today = date.today()
        this_month_num, this_year = today.month, today.year
        today_sales, month_sales, today_purchase, month_purchase = 0.0, 0.0, 0.0, 0.0
        monthly_trend, top_clients, doc_counts = {}, {}, {"invoice": 0, "cn": 0, "po": 0, "grn": 0, "bill": 0, "dn": 0}

        for inv in invoices:
            cat, dtype, is_cn = inv.get('doc_category', 'sale'), inv.get('doc_type', 'invoice'), inv.get('is_credit_note', False)
            if inv.get('status') == 'Cancelled': continue
            amt = float(inv.get('grand_total', 0))

            try:
                parts = inv.get('invoice_date', '').split('-')
                if len(parts) == 3:
                    m = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
                    inv_date = date(int(parts[2]), m.get(parts[1], 1), int(parts[0]))
                    inv_month_num, inv_year, inv_month_label, inv_today = inv_date.month, inv_date.year, inv_date.strftime('%b %y'), (inv_date == today)
                else: raise ValueError
            except: inv_date, inv_month_num, inv_year, inv_month_label, inv_today = None, 0, 0, '', False

            if is_cn: doc_counts['cn'] = doc_counts.get('cn', 0) + 1
            elif dtype in doc_counts: doc_counts[dtype] = doc_counts.get(dtype, 0) + 1

            if cat == 'sale' and not is_cn and dtype == 'invoice':
                if inv_today: today_sales += amt
                if inv_month_num == this_month_num and inv_year == this_year: month_sales += amt
                if inv_date: monthly_trend[inv_month_label] = monthly_trend.get(inv_month_label, 0) + amt
                cname = inv.get('client_name', 'Unknown')
                top_clients[cname] = top_clients.get(cname, 0) + amt
            elif cat == 'purchase' and dtype == 'bill':
                if inv_today: today_purchase += amt
                if inv_month_num == this_month_num and inv_year == this_year: month_purchase += amt

        try:
            out_res = outstanding_report().get_json()
            out_count, out_tot = (len(out_res), sum(o['balance'] for o in out_res)) if isinstance(out_res, list) else (0, 0)
        except: out_count, out_tot = 0, 0

        try:
            res = supabase.table('inventory_products').select('data').eq('tenant_id', get_tenant_id()).execute()
            low_stock = []
            for r in res.data:
                item = r['data']
                stock, reorder = float(item.get('current_stock', 0)), float(item.get('reorder_level', 0))
                if stock <= reorder: low_stock.append({"item": item.get('item_name', ''), "stock": stock, "reorder": reorder})
        except: low_stock = []

        return jsonify({
            "today_sales": round(today_sales, 2), "month_sales": round(month_sales, 2),
            "today_purchase": round(today_purchase, 2), "month_purchase": round(month_purchase, 2),
            "outstanding_count": out_count, "outstanding_total": round(out_tot, 2),
            "low_stock_count": len(low_stock), "low_stock_items": low_stock[:5],
            "monthly_trend": sorted(monthly_trend.items(), key=lambda x: datetime.strptime(x[0], '%b %y'))[-6:],
            "top_clients": sorted(top_clients.items(), key=lambda x: x[1], reverse=True)[:5],
            "doc_counts": doc_counts, "total_invoices": len(invoices)
        })
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard(): return render_template('dashboard.html')

@app.route('/reset-password', methods=['POST'])
@login_required
def reset_password():
    if not current_user.is_master: return "Unauthorized", 403
    target, new_pass = request.form.get('target_user_id'), request.form.get('reset_password')
    if target and new_pass:
        try:
            supabase.table('app_users').update({"password": generate_password_hash(new_pass)}).eq('username', target).execute()
            flash(f"Password updated successfully!", "success")
        except Exception as e: flash(f"Database error: {e}", "error")
    return redirect(url_for('user_profile', edit_user=target))

if __name__ == "__main__":
    # Check if we are running locally in development mode
    # Set FLASK_ENV=development in your local .env to turn on debugging locally
    is_dev = os.getenv("FLASK_ENV") == "development"
    
    # Run the app securely
    app.run(debug=is_dev, host="0.0.0.0", port=int(os.getenv("PORT", 5000)))